import openpyxl
import datetime

excel_path = r"C:\Users\user\Desktop\Antigravity\データ分析自動化\スタジアム_データ.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['【AI】予想・答え合わせ']

print("=== Predictions Non-Empty Rows (Last 15) ===")
non_empty_rows = []
for r in range(4, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, 9)]
    if any(v is not None for v in vals):
        non_empty_rows.append((r, vals))
        
print(f"Total non-empty rows: {len(non_empty_rows)}")
for r_num, vals in non_empty_rows[-20:]:
    print(f"Row {r_num}: {vals}")

wb.close()
