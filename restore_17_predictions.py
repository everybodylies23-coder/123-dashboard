import openpyxl
import os
import datetime
import re

dir_path = r"C:\Users\user\Desktop\Antigravity\データ分析自動化"
excel_path = os.path.join(dir_path, "スタジアム_データ.xlsx")
report_path = os.path.join(dir_path, "data_input", "スタジアム　20260716_ai_report.md")

def parse_report_predictions():
    if not os.path.exists(report_path):
        print(f"Error: Report file {report_path} not found.")
        return []
        
    print(f"Parsing predictions from {report_path}...")
    with open(report_path, "r", encoding="utf-8") as f:
        content = f.read()
        
    table_pattern = r'\|?\s*(?:\d{4}/\d{2}/\d{2})\s*\|.*'
    table_lines = re.findall(table_pattern, content)
    
    predictions = []
    for line in table_lines:
        cells = [c.strip() for c in line.split("|")]
        if len(cells) > 0 and cells[0] == "":
            cells.pop(0)
        if len(cells) > 0 and cells[-1] == "":
            cells.pop()
            
        if len(cells) >= 4:
            date_str = cells[0]
            machine_name = cells[1]
            try:
                machine_num = int(cells[2])
            except ValueError:
                machine_num = cells[2]
            reason = cells[3]
            
            if "機種名" in machine_name or "---" in machine_name:
                continue
                
            predictions.append({
                "date_str": date_str,
                "machine_name": machine_name,
                "machine_number": machine_num,
                "reason": reason
            })
            
    print(f"Parsed {len(predictions)} predictions for target date {predictions[0]['date_str'] if predictions else 'N/A'}")
    return predictions

def insert_and_calculate():
    preds_to_insert = parse_report_predictions()
    if not preds_to_insert:
        print("No predictions parsed. Aborting.")
        return
        
    print(f"Opening workbook: {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ai_ws = wb['【AI】予想・答え合わせ']
    
    # Let's find where to insert. We look for the first row after July 16th.
    insert_row_idx = None
    for r in range(4, ai_ws.max_row + 1):
        dt_val = ai_ws.cell(r, 1).value
        if dt_val:
            if isinstance(dt_val, str):
                try:
                    dt_val = datetime.datetime.strptime(dt_val.split()[0].replace("-", "/"), "%Y/%m/%d")
                except Exception:
                    pass
            if isinstance(dt_val, (datetime.datetime, datetime.date)):
                # If date is >= July 18th, insert here!
                if dt_val.strftime("%Y/%m/%d") >= "2026/07/18":
                    insert_row_idx = r
                    break
                    
    if not insert_row_idx:
        # If no row >= July 18th is found, just append to the end of non-empty rows
        for r in range(ai_ws.max_row, 3, -1):
            if ai_ws.cell(r, 1).value is not None:
                insert_row_idx = r + 1
                break
        if not insert_row_idx:
            insert_row_idx = 4
            
    print(f"Inserting {len(preds_to_insert)} rows at row index {insert_row_idx}...")
    ai_ws.insert_rows(insert_row_idx, len(preds_to_insert))
    
    # Write the predictions
    for i, p in enumerate(preds_to_insert):
        curr_r = insert_row_idx + i
        dt_val = datetime.datetime.strptime(p["date_str"], "%Y/%m/%d")
        
        dt_cell = ai_ws.cell(curr_r, 1, dt_val)
        dt_cell.number_format = 'yyyy/mm/dd'
        ai_ws.cell(curr_r, 2, p["machine_name"])
        ai_ws.cell(curr_r, 3, p["machine_number"])
        ai_ws.cell(curr_r, 4, p["reason"])
        
    # Copy formulas down to Columns 5-8 for the new rows and shifted rows
    # Formula source is Row 4
    for r in range(insert_row_idx, ai_ws.max_row + 1):
        # We only apply formulas if Col 1 is not empty
        if ai_ws.cell(r, 1).value is not None:
            # Shift formulas in E-H
            # E
            formula_e = f"=IF(OR(A{r}=\"\",C{r}=\"\"),\"\",SUMIFS('【データ】蓄積用'!$D$2:$D$15000,'【データ】蓄積用'!$A$2:$A$15000,A{r},'【データ】蓄積用'!$C$2:$C$15000,C{r}))"
            ai_ws.cell(r, 5, formula_e)
            # F
            formula_f = f"=IF(OR(A{r}=\"\",C{r}=\"\"),\"\",SUMIFS('【データ】蓄積用'!$E$2:$E$15000,'【データ】蓄積用'!$A$2:$A$15000,A{r},'【データ】蓄積用'!$C$2:$C$15000,C{r}))"
            ai_ws.cell(r, 6, formula_f)
            # G
            formula_g = f"=IF(OR(A{r}=\"\",C{r}=\"\"),\"\",SUMIFS('【データ】蓄積用'!$L$2:$L$15000,'【データ】蓄積用'!$A$2:$A$15000,A{r},'【データ】蓄積用'!$C$2:$C$15000,C{r}))"
            ai_ws.cell(r, 7, formula_g)
            # H
            formula_h = f"=IF(A{r}=\"\",\"\",IF(OR(E{r}=\"\",E{r}=\"集計\"),\"\",IF(ISNUMBER(SEARCH(\"ジャグラー\",B{r})),IF(AND(IFERROR(VALUE(G{r}),0)>=4.5,IFERROR(VALUE(F{r}),0)>=500),\"〇\",\"×\"),IF(IFERROR(VALUE(F{r}),0)>=1000,\"〇\",\"×\"))))"
            ai_ws.cell(r, 8, formula_h)

    print("Formulas written. Now performing pre-calculated lookup...")
    
    # Now run the lookup calculations as static values to speed up and avoid errors
    data_ws = wb['【データ】蓄積用']
    accumulated_db = {}
    for r in range(2, data_ws.max_row + 1):
        date_val = data_ws.cell(r, 1).value
        mach_num = data_ws.cell(r, 3).value
        g_games = data_ws.cell(r, 4).value
        diff_coins = data_ws.cell(r, 5).value
        setting_score = data_ws.cell(r, 29).value  # AC (Col 29)
        
        if date_val is not None:
            # normalize date to match
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime("%Y/%m/%d")
            elif isinstance(date_val, datetime.date):
                date_str = date_val.strftime("%Y/%m/%d")
            else:
                date_str = str(date_val).strip().replace("-", "/")
            try:
                m_num = int(str(mach_num).strip())
                accumulated_db[(date_str, m_num)] = (g_games, diff_coins, setting_score)
            except ValueError:
                continue
                
    rewritten_count = 0
    for r in range(4, ai_ws.max_row + 1):
        date_val = ai_ws.cell(r, 1).value
        mach_val = ai_ws.cell(r, 3).value
        m_name = str(ai_ws.cell(r, 2).value or "")
        
        if date_val is not None and mach_val is not None:
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime("%Y/%m/%d")
            elif isinstance(date_val, datetime.date):
                date_str = date_val.strftime("%Y/%m/%d")
            else:
                date_str = str(date_val).strip().replace("-", "/")
                
            try:
                m_num = int(str(mach_val).strip())
                key = (date_str, m_num)
                if key in accumulated_db:
                    g_games, diff_coins, setting_score = accumulated_db[key]
                    
                    try:
                        coins_int = int(str(diff_coins).replace(",", "").replace("+", "").strip())
                    except ValueError:
                        coins_int = 0
                        
                    score_val = 0
                    if setting_score is not None and setting_score != "":
                        try:
                            score_val = float(str(setting_score).strip())
                        except ValueError:
                            pass
                            
                    # Juggler: score >= 4.5 AND diff_coins >= 500
                    # Others: diff_coins >= 1000
                    if "ジャグラー" in m_name:
                        status = "〇" if (score_val >= 4.5 and coins_int >= 500) else "×"
                    else:
                        status = "〇" if coins_int >= 1000 else "×"
                        
                    ai_ws.cell(r, 5, g_games)
                    ai_ws.cell(r, 6, diff_coins)
                    ai_ws.cell(r, 7, score_val if setting_score is not None else "")
                    ai_ws.cell(r, 8, status)
                    rewritten_count += 1
            except ValueError:
                continue
                
    print(f"Saving workbook (populated {rewritten_count} answer rows)...")
    while True:
        try:
            wb.save(excel_path)
            break
        except PermissionError:
            print("Excel file is open. Please close it and press ENTER to retry...")
            input()
            
    wb.close()
    print("Done restoring July 17th predictions.")

if __name__ == "__main__":
    insert_and_calculate()
    
    # Re-generate the HTML dashboard and push to GitHub Pages
    import auto_analyzer
    auto_analyzer.generate_html_dashboard(excel_path, "スタジアム")
    
    # Deploy to GitHub Pages
    print("\nDeploying to GitHub Pages...")
    # Find git.exe
    git_path = None
    for p in [r"C:\Program Files\Git\cmd", r"C:\Program Files (x86)\Git\cmd", os.path.expandvars(r"%LOCALAPPDATA%\Programs\Git\cmd")]:
        if os.path.exists(os.path.join(p, "git.exe")):
            git_path = p
            break
            
    if git_path:
        os.environ["PATH"] = git_path + os.pathsep + os.environ["PATH"]
        
    os.system("git add .")
    os.system('git commit -m "Restore July 17th predictions and update answers"')
    res = os.system("git push origin main")
    if res != 0:
        os.system("git push origin master")
    print("GitHub Pages deployment finished.")
