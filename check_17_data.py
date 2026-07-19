import openpyxl
import datetime

excel_path = r"C:\Users\user\Desktop\Antigravity\データ分析自動化\スタジアム_データ.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Check accumulation max date
ws_data = wb['【データ】蓄積用']
dates_data = []
for r in range(2, ws_data.max_row + 1):
    val = ws_data.cell(r, 1).value
    if isinstance(val, (datetime.datetime, datetime.date)):
        dates_data.append(val)
print("【データ】蓄積用:")
if dates_data:
    print(f"  Count: {len(dates_data)}")
    print(f"  Min date: {min(dates_data)}")
    print(f"  Max date: {max(dates_data)}")
else:
    print("  No dates found.")

# Check predictions sheet dates
ws_pred = wb['【AI】予想・答え合わせ']
dates_pred = []
for r in range(4, ws_pred.max_row + 1):
    val = ws_pred.cell(r, 1).value
    if isinstance(val, (datetime.datetime, datetime.date)):
        dates_pred.append(val)
print("\n【AI】予想・答え合わせ:")
if dates_pred:
    print(f"  Count: {len(dates_pred)}")
    print(f"  Min date: {min(dates_pred)}")
    print(f"  Max date: {max(dates_pred)}")
    
    # Show rows around the end
    print("  Last 10 rows:")
    for r in range(max(4, ws_pred.max_row - 10), ws_pred.max_row + 1):
        row_vals = [ws_pred.cell(r, c).value for c in range(1, 9)]
        print(f"    Row {r}: {row_vals}")
else:
    print("  No dates found.")
    
# Let's check other xlsx files in the directory for July 17th predictions
import os
for f in ["スタジアム_データ_corrupt_backup.xlsx", "スタジアム_データ_軽量版.xlsx", "データ分析_テンプレート_v13.11.xlsx"]:
    path = os.path.join(r"C:\Users\user\Desktop\Antigravity\データ分析自動化", f)
    if os.path.exists(path):
        try:
            wb_other = openpyxl.load_workbook(path, data_only=True)
            for sheetname in wb_other.sheetnames:
                if "予想" in sheetname:
                    ws_other = wb_other[sheetname]
                    other_dates = []
                    for r in range(4, ws_other.max_row + 1):
                        val = ws_other.cell(r, 1).value
                        if isinstance(val, (datetime.datetime, datetime.date)):
                            other_dates.append(val)
                    if other_dates:
                        print(f"\nFile {f} - {sheetname}: max date = {max(other_dates)}, count = {len(other_dates)}")
                        if max(other_dates) >= datetime.datetime(2026, 7, 16):
                            # print last 5 rows
                            print("  Last 5 rows in other:")
                            for r in range(max(4, ws_other.max_row - 5), ws_other.max_row + 1):
                                row_vals = [ws_other.cell(r, c).value for c in range(1, 9)]
                                if any(row_vals):
                                    print(f"    Row {r}: {row_vals}")
            wb_other.close()
        except Exception as e:
            print(f"Error checking {f}: {e}")

wb.close()
